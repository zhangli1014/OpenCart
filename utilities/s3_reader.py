import boto3
from openpyxl import load_workbook
from io import BytesIO

def get_data_s3(bucket_name,file_key):
    s3 = boto3.client('s3')
    try:
        response = s3.get_object(Bucket=bucket_name, Key=file_key)
        file_content = response['Body'].read()
        workbook = load_workbook(filename=BytesIO(file_content))
        return workbook
    except Exception as e:
        print(f"Failed to load data: {e}")
        return None

def getRowCount_froms3(workbook,sheetname):
    sheet = workbook[sheetname]
    return sheet.max_row

def getColumnCount_froms3(workbook,sheetname):
    #workbook = get_data_s3(bucket_name,file_key)
    sheet = workbook[sheetname]
    return sheet.max_column

def readData_froms3(workbook,sheetname,rownum,columnno):
    #workbook = get_data_s3(bucket_name,file_key)
    sheet = workbook[sheetname]
    return sheet.cell(row=rownum,column=columnno).value

if __name__=='__main__':
    print(get_data_s3('opencart-login-testdata','LoginInfo.json'))